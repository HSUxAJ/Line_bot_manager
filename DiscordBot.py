import discord
from discord.ext import commands
import json
from linebot import LineBotApi, WebhookHandler
from linebot.models import *
import os
from dotenv import load_dotenv
from openpyxl import load_workbook

intents = discord.Intents.default()
intents.webhooks = True
intents.guilds = True
intents.webhooks = True
intents.message_content = True
bot = commands.Bot(command_prefix='!', intents=intents)

load_dotenv()
line_bot_api = LineBotApi(os.getenv('LINE_CHANNEL_ACCESS_TOKEN', None))
discord_token = os.getenv('DISCORD_TOKEN', None)
line_group_id = {}

@bot.event
async def on_message(message):
	if message.author.bot: 	return
	channel_id = str(message.channel.id)
	if channel_id == '1147078580266860646': #上傳任務區
		await load_excel(message)
	else:
		bot_send_message(message, channel_id)

async def load_excel(message):
	attachment = message.attachments[0]
	file_data = await attachment.read()

	# 將二進制數據保存到本地文件
	with open("temp_excel.xlsx", "wb") as temp_file:
			temp_file.write(file_data)

	# 打開 Excel 文件並解析
	workbook = load_workbook("temp_excel.xlsx")
	sheet = workbook.active
	# 获取表头，即第一行的数据
	headers = []
	for cell in sheet[1]:
		if cell.value != 'None':
			headers.append(cell.value)

	# 遍历余下的行并将其转换为字典
	data = []
	for row in sheet.iter_rows(min_row=2, values_only=True):
		row_dict = {header: value for header, value in zip(headers, row)}
		data.append(row_dict)
	print(data[0])

def bot_send_message(message, channel_id):
	if channel_id not in line_group_id:
		update()
	if message.attachments == []:
		line_bot_api.push_message(line_group_id[channel_id]['line_group_id'], TextSendMessage(text=message.content))
	else:
		image_url = str(message.attachments[0])
		image_message = ImageSendMessage(
			original_content_url=image_url,
			preview_image_url=image_url
			)
		line_bot_api.push_message(line_group_id[channel_id]['line_group_id'], image_message)
	
def update():
	global line_group_id
	with open('data.json', 'r', encoding='utf-8') as f:
	    line_group_id = json.load(f)

def run_bot():
    update()
    bot.run(discord_token)

if __name__ == '__main__':
    run_bot()


#