import os
import sys
from argparse import ArgumentParser
from dotenv import load_dotenv

import asyncio
import aiohttp
from aiohttp import web

import logging

from aiohttp.web_runner import TCPSite

from linebot import (
    AsyncLineBotApi, WebhookParser
)
from linebot.aiohttp_async_http_client import AiohttpAsyncHttpClient
from linebot.exceptions import (
    InvalidSignatureError
)
from linebot.models import (
    MessageEvent, TextMessage, TextSendMessage,
)

from linebot import LineBotApi, WebhookHandler
from linebot.exceptions import InvalidSignatureError
from linebot.models import MessageEvent, TextMessage, TextSendMessage

import multiprocessing
import DiscordBot
import requests
import json

from apscheduler.schedulers.background import BackgroundScheduler
from datetime import datetime, timedelta

import random
import string

from openpyxl import load_workbook

load_dotenv()
channel_secret = os.getenv('LINE_CHANNEL_SECRET', None)
channel_access_token = os.getenv('LINE_CHANNEL_ACCESS_TOKEN', None)
if channel_secret is None:
    print('Specify LINE_CHANNEL_SECRET as environment variable.')
    sys.exit(1)
if channel_access_token is None:
    print('Specify LINE_CHANNEL_ACCESS_TOKEN as environment variable.')
    sys.exit(1)

class Handler:
    def __init__(self, line_bot_api, parser, line):
        self.line_bot_api = line_bot_api
        self.line = line
        self.parser = parser
        self.message_history = {}
        with open('data.json', 'r', encoding='utf-8') as f:
            self.discord_info = json.load(f)
        self.scheduler = BackgroundScheduler()
        
    async def callback(self, request):
        signature = request.headers['X-Line-Signature']
        body = await request.text()

        try:
            events = self.parser.parse(body, signature)
        except InvalidSignatureError:
            return web.Response(status=400, text='Invalid signature')

        for event in events:
            if event.source.type != 'group':
                return
            
            group_id = event.source.group_id
            user_id = event.source.user_id
            group_summary = await self.line_bot_api.get_group_summary(group_id)
            group_name = group_summary.group_name
            pwds = list(self.discord_info['manager_id'].keys()) #所有密碼
            no_used_sub_pwds = [] #沒使用過的子群密碼
            no_used_m_pwds = [] #沒使用過的管理群密碼
            no_used_response_pwds = [] #沒使用過的回傳群密碼
            for p in pwds:
                if self.discord_info['manager_id'][p] == ["", "", {}, []]: # data(manager_id)預設樣式!
                    no_used_m_pwds.append(p)
            t = list(self.discord_info['manager_id'].values())
            ids = []
            mana_id = []
            for tt in t:
                if (tt[1] != ''):
                    ids.append(tt[1])
                if (tt[0] != ''):
                    mana_id.append(tt[0])
            for data in self.discord_info['manager_id'].values(): #取得子群資訊
                if data[2] != {}:
                    no_used_response_pwds.append(list(data[2].keys())[0])
                    pwds.append(list(data[2].keys())[0])
                for d in data[3]:
                    for k in d:
                        pwds.append(k)
                        if d[k] != '':
                            ids.append(d[k])
                        else:
                            no_used_sub_pwds.append(k)


            if event.message.type == "text":
                content_message = event.message.text
                if content_message == "new子群":
                    new_pwd = self.generate_random_string(20)
                    for val in list(self.discord_info['manager_id'].values()):
                        if group_id == val[1]:
                            if len(val[3]) < 10: #子群最多10個
                                val[3].append({new_pwd: ''})
                            else:
                                self.line.push_message(group_id, TextMessage(text='子群數量已達上限!'))   
                                return
                    self.line.push_message(group_id, TextMessage(text='子群啟用密碼為：' + new_pwd))   
                    self.update_data()
                elif content_message == "new回報區":
                    new_pwd = self.generate_random_string(20)
                    for val in list(self.discord_info['manager_id'].values()):
                        print(val[2])
                        if group_id == val[1]:
                            if val[2] == {}:
                                val[2] = {new_pwd: ""}
                            else:
                                self.line.push_message(group_id, TextMessage(text='回報區已存在!'))   
                                return
                    self.line.push_message(group_id, TextMessage(text='回報區啟用密碼為：' + new_pwd))   
                    self.update_data()                    
                elif content_message in pwds:
                    pw = content_message
                    if group_id in ids:
                        self.line.push_message(group_id, TextMessage(text='群組不需再次啟用~'))
                    elif pw in no_used_m_pwds:
                        self.discord_info['manager_id'][pw] = [user_id, group_id, {}, []]
                        self.line.push_message(group_id, TextMessage(text='管理群組成功啟用~'))
                    elif pw in no_used_sub_pwds:
                        for data in self.discord_info['manager_id'].values():
                            for d in data[3]:
                                for k in d:
                                    if k == pw:
                                        d[k] = group_id
                                        self.line.push_message(group_id, TextMessage(text=f'子群組成功啟用~({len(data[3])}/10)'))
                                        break
                    elif pw in no_used_response_pwds:
                        for data in self.discord_info['manager_id'].values():
                            if pw == list(data[2].keys())[0]:
                                data[2] = {pw: group_id}
                                self.line.push_message(group_id, TextMessage(text='回報區群組成功啟用~'))
                                break
                    else:
                        self.line.push_message(group_id, TextMessage(text='此密碼已被使用過~'))
                    self.update_data()

            elif event.message.type == "location":
                content_message = event.message.address
            
            elif event.message.type == 'file':
                ok = False
                tar = []
                for arr in self.discord_info['manager_id'].values():
                    if arr[0] == user_id and arr[1] == group_id:
                        ok = True
                        tar = arr
                        break
                if ok:
                    file_content = await self.line_bot_api.get_message_content(event.message.id)
                    file_path = os.path.join(os.path.dirname(__file__), 'sc.xlsx')
                    with open(file_path, 'wb') as f:
                        async for chunk in file_content.iter_content():
                            f.write(chunk)
                    await self.load_schedule(group_id, tar[3])
                    self.clear_excel(file_path)
                else:
                    self.line.push_message(group_id, TextMessage(text='此群組權限不足!'))
                return
            
            if group_id not in self.discord_info:
                new_info = self.create_discord_channel(line_group_id=group_id, group_name=group_name)
                self.discord_info.update(new_info)
                self.update_data()

            request_data = await Handler.create_request_data(self.line_bot_api, group_id, user_id, content_message)
            requests.post(url=self.discord_info[group_id]['webhook'], data=request_data)
            
            if group_id not in self.message_history:
                self.message_history[group_id] = []
            self.message_history[group_id].append([event.message.type, content_message, datetime.now()]) #歷史訊息儲存方式[型態，內容，時間]
        return web.Response(text="OK\n")
    
    def generate_random_string(self, length):
        characters = string.digits + string.ascii_lowercase + string.ascii_uppercase
        random_string = ''.join(random.sample(characters, length))
        return random_string
    
    def update_data(self):
        with open('data.json', 'w', encoding='utf-8') as f:
            json.dump(self.discord_info, f, ensure_ascii=False)
    
    def create_discord_channel(self, line_group_id, group_name):
        discord_token = os.getenv('DISCORD_TOKEN', None)
        discord_guild_id = os.getenv('DISCORD_GUILD_ID', None)
        api_endpoint_channels = f'https://discord.com/api/v10/guilds/{discord_guild_id}/channels'

        headers = {
            'Authorization': f'Bot {discord_token}',
            'Content-Type': 'application/json',
        }

        # 創建文字頻道
        channel_data = {
            'name': group_name,  # 頻道名稱
            'type': 0,  # 文字頻道的類型
        }

        response_channel = requests.post(api_endpoint_channels, json=channel_data, headers=headers)
        if response_channel.status_code == 201:
            new_channel_data = response_channel.json()
            print(f'文字頻道 {new_channel_data["name"]} 已成功創建！')
        else:
            print('創建文字頻道時出現問題。')

        # 創建 Webhook
        API_ENDPOINT = f'https://discord.com/api/v10/channels/{new_channel_data["id"]}/webhooks'
        headers = {
            'Authorization': f'Bot {discord_token}',
            'Content-Type': 'application/json',
        }

        data = {
            'name': 'Webhook',  # Webhook 名稱
        }

        response = requests.post(API_ENDPOINT, json=data, headers=headers)
        if response.status_code == 200:
            new_webhook_data = response.json()
            print(f'Webhook 已成功創建，Webhook ID：{new_webhook_data["id"]}')
        else:
            print('創建 Webhook 時出現問題。')

        ret = {}
        ret[line_group_id] = {
            'name': group_name,
            'webhook': new_webhook_data["url"]
        }
        ret[new_channel_data["id"]] = {'line_group_id': line_group_id}
        return ret

    async def create_request_data(self, group_id, user_id, text=None):
        profile = await self.get_group_member_profile(group_id, user_id)
        
        headers = {
            "content-type": "application/json; charset=UTF-8",
            "Authorization": "Bearer {}".format(os.environ['LINE_CHANNEL_ACCESS_TOKEN'])
        }
        url = 'https://api.line.me/v2/bot/group/' + group_id + '/summary'

        async with aiohttp.ClientSession() as session:
            async with session.get(url, headers=headers) as response:
                response_json = await response.json()

        request_data = {
            "content": text,
            "username": profile.display_name,
            "avatar_url": profile.picture_url
        }
        return request_data

    def check_message(self, id):
        print('Time for checking!!')
        now = datetime.now()
        content_message = ""
        if id in self.message_history:
            if self.message_history[id][-1][0] == 'bot':
                cnt = 0
                for data in self.message_history[id][::-1]:
                    if data[0] != 'bot':
                        break
                    cnt += 1
                if cnt < 5: #最多重複傳幾次
                    self.line.push_message(id, TextSendMessage(text=self.message_history[id][-1][1]))
                    self.message_history[id].append(['bot', self.message_history[id][-1][1], datetime.now()])
                    self.scheduler.add_job(self.check_message, 'date', [id], run_date=datetime.now()+timedelta(seconds=10)) #檢查回復的間隔
                    return
                else:
                    content_message = '" ' + self.discord_info[id]['name'] + " \" didn't make any response!!"
            # elif self.message_history[id][-1][0] == 'location' and (now - self.message_history[id][-1][2]).total_seconds() < 10: #限制回傳與提醒時間間距
            #     content_message = '" ' + self.discord_info[id]['name'] + ' " located at ' + self.message_history[id][-1][1]
        
        #回報地點
        dic = self.discord_info['manager_id']
        mana_id = ''
        for mid in dic:
            for val in dic[mid][3]:
                if id == list(val.values())[0]:
                    mana_id = list(dic[mid][2].values())[0]
                    break
        self.line.push_message(mana_id, TextSendMessage(text=content_message))
        
    async def load_schedule(self, mana_id, tar):
        info = []
        for v in tar:
            id = list(v.values())[0]
            if id == "":
                continue
            group_summary = await self.line_bot_api.get_group_summary(id)
            group_name = group_summary.group_name
            info.append([group_name, id])

        workbook = load_workbook("sc.xlsx")
        sheet = workbook.active
        headers = []
        for cell in sheet[1]:
            if cell.value != 'None':
                headers.append(cell.value)

        data = []
        for row in sheet.iter_rows(min_row=2, values_only=True):
            row_dict = {header: value for header, value in zip(headers, row)}
            data.append(row_dict)

        for event in data:
            if event['name'] == None:
                continue
            id = ''
            find = False
            for i in info:
                if i[0] == event['name']:
                    id = i[1]
                    find = True
                    break
            if not find:
                self.line.push_message(mana_id, TextMessage(text=str(event['name']) + '群組不是子群，發送失敗'))
                continue
            def push_msg(id, task):
                self.line.push_message(id, TextSendMessage(text=task))
                content = ['bot', task, datetime.now()]
                if id not in self.message_history:
                    self.message_history[id] = []
                self.message_history[id].append(content)
            self.scheduler.add_job(push_msg, 'date', [id, event['task']], run_date = event['time'])
            self.scheduler.add_job(self.check_message, 'date', [id], run_date = event['time'] + timedelta(seconds=10)) #檢查回復的間隔
            # time = datetime.now() + timedelta(seconds=10)
            # self.scheduler.add_job(push_msg, 'date', [id, event['task']], run_date=time)
            # self.scheduler.add_job(self.check_message, 'date', [id], run_date=time+timedelta(seconds=10))

    def clear_excel(self, file_path):
        try:
            # 加载 Excel 文件
            wb = load_workbook(file_path)
            # 清空所有的工作表
            for sheet_name in wb.sheetnames:
                sheet = wb[sheet_name]
                for row in sheet.iter_rows():
                    for cell in row:
                        cell.value = None
            # 保存修改后的 Excel 文件
            wb.save(file_path)
            print(f"Excel file at {file_path} has been cleared successfully.")
        except Exception as e:
            print(f"Error clearing Excel file: {e}")

async def main(port=8080):
    session = aiohttp.ClientSession()
    async_http_client = AiohttpAsyncHttpClient(session)
    line_bot_api = AsyncLineBotApi(channel_access_token, async_http_client)
    parser = WebhookParser(channel_secret)
    line = LineBotApi(channel_access_token)

    handler = Handler(line_bot_api, parser, line)

    app = web.Application()
    app.add_routes([web.post('/callback', handler.callback)])

    runner = web.AppRunner(app)
    await runner.setup()
    site = TCPSite(runner=runner, port=port)
    await site.start()

    handler.scheduler.start()

    while True:
        await asyncio.sleep(3600)  # sleep forever

if __name__ == "__main__":
    p = multiprocessing.Process(target=DiscordBot.run_bot)
    p.start()
    logging.basicConfig(format='%(levelname)s:%(message)s', level=logging.INFO)

    arg_parser = ArgumentParser(
        usage='Usage: python ' + __file__ + ' [--port <port>] [--help]'
    )
    arg_parser.add_argument('-p', '--port', type=int, default=8080, help='port')
    options = arg_parser.parse_args()

    asyncio.run(main(options.port))